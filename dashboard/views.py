import io
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.views import LoginView
from django.db.models import Avg, Count, F, Q, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DeleteView, DetailView, ListView, TemplateView, UpdateView

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from accounts.models import Address, User
from catalog.forms import CategoryForm, HeroBannerForm, ProductForm
from catalog.models import Category, Collection, HeroBanner, Product, ProductImage, Wishlist
from coupons.models import Coupon
from dashboard.models import AuditLog, LoginAttempt, StoreSetting
from orders.models import Order, OrderItem, ReturnRequest
from payments.models import Payment
from reviews.models import ProductQuestion, Review


class StaffLoginView(LoginView):
    """Staff Login Page for Super Admin."""
    template_name = "dashboard/login.html"

    def get_success_url(self):
        return reverse("dashboard:home")


class StaffLogoutView(View):
    """Unified Logout for Super Admin."""
    def get(self, request):
        logout(request)
        messages.success(request, "Super Admin logged out successfully.")
        return redirect("dashboard:login")

    def post(self, request):
        return self.get(request)


class SuperAdminRequiredMixin(UserPassesTestMixin):
    """
    Protects every Super Admin route.
    Only allows users with is_staff=True or is_superuser=True.
    """
    login_url = "dashboard:login"

    def test_func(self):
        request = getattr(self, "request", None)
        user = request.user if request else None
        return bool(user and user.is_authenticated and (user.is_staff or user.is_superuser))

    def handle_no_permission(self):
        request = getattr(self, "request", None)
        if not request or not request.user.is_authenticated:
            path = request.path if request else ""
            return redirect(f"{reverse('dashboard:login')}?next={path}")
        return render(request, "403.html", status=403)


# ---------------------------------------------------------------------------
# 1. Dashboard Overview
# ---------------------------------------------------------------------------

class DashboardHomeView(SuperAdminRequiredMixin, TemplateView):
    template_name = "dashboard/home.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        now = timezone.now()
        today = now.date()
        this_month_start = today.replace(day=1)
        this_year_start = today.replace(month=1, day=1)

        # 11 Key Performance Indicator (KPI) Cards
        ctx["total_products"] = Product.objects.count()
        ctx["total_customers"] = User.objects.filter(is_staff=False).count()

        valid_order_statuses = [
            Order.Status.CONFIRMED,
            Order.Status.PROCESSING,
            Order.Status.SHIPPED,
            Order.Status.OUT_FOR_DELIVERY,
            Order.Status.DELIVERED,
        ]

        ctx["orders_today"] = Order.objects.filter(placed_at__date=today).count()
        ctx["pending_orders"] = Order.objects.filter(status=Order.Status.PENDING).count()
        ctx["completed_orders"] = Order.objects.filter(status=Order.Status.DELIVERED).count()
        ctx["cancelled_orders"] = Order.objects.filter(status=Order.Status.CANCELLED).count()

        ctx["revenue_today"] = Order.objects.filter(
            status__in=valid_order_statuses, placed_at__date=today
        ).aggregate(total=Sum("grand_total"))["total"] or Decimal("0.00")

        ctx["monthly_revenue"] = Order.objects.filter(
            status__in=valid_order_statuses, placed_at__date__gte=this_month_start
        ).aggregate(total=Sum("grand_total"))["total"] or Decimal("0.00")

        ctx["yearly_revenue"] = Order.objects.filter(
            status__in=valid_order_statuses, placed_at__date__gte=this_year_start
        ).aggregate(total=Sum("grand_total"))["total"] or Decimal("0.00")

        low_stock_qs = Product.objects.filter(stock_quantity__lte=5, stock_quantity__gt=0)
        out_of_stock_qs = Product.objects.filter(stock_quantity=0)

        ctx["low_stock_count"] = low_stock_qs.count()
        ctx["out_of_stock_count"] = out_of_stock_qs.count()
        ctx["low_stock_products"] = low_stock_qs.select_related("category")[:5]

        # Quick Tables
        ctx["recent_orders"] = Order.objects.all().select_related("user").order_by("-placed_at")[:8]
        ctx["recent_customers"] = User.objects.filter(is_staff=False).order_by("-date_joined")[:5]
        ctx["recent_reviews"] = Review.objects.all().select_related("product", "user").order_by("-created_at")[:5]
        ctx["active_coupons"] = Coupon.objects.filter(is_active=True)[:5]

        # Chart.js Analytics Data
        ctx["category_sales"] = list(
            Category.objects.annotate(
                total_products=Count("products"),
                total_sales=Count("products__orderitem")
            ).values("name", "total_products", "total_sales").order_by("-total_sales")[:6]
        )

        ctx["payment_methods_stat"] = list(
            Payment.objects.values("gateway").annotate(count=Count("id"), total=Sum("amount"))
        )

        ctx["top_selling_products"] = Product.objects.annotate(
            total_sold=Sum("orderitem__quantity")
        ).filter(total_sold__gt=0).order_by("-total_sold")[:5]

        return ctx


# ---------------------------------------------------------------------------
# 2. Product Management (Products, Categories, Collections, Inventory, Reviews)
# ---------------------------------------------------------------------------

class ProductManagementView(SuperAdminRequiredMixin, ListView):
    template_name = "dashboard/products.html"
    context_object_name = "products"
    paginate_by = 15

    def get_queryset(self):
        qs = Product.objects.all().select_related("category").prefetch_related("images")
        q = self.request.GET.get("q", "").strip()
        category = self.request.GET.get("category", "")
        status = self.request.GET.get("status", "")

        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(sku__icontains=q))
        if category:
            qs = qs.filter(category_id=category)
        if status == "published":
            qs = qs.filter(is_published=True)
        elif status == "draft":
            qs = qs.filter(is_published=False)
        elif status == "low_stock":
            qs = qs.filter(stock_quantity__lte=5, stock_quantity__gt=0)
        elif status == "out_of_stock":
            qs = qs.filter(stock_quantity=0)

        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["categories"] = Category.objects.all()
        ctx["q"] = self.request.GET.get("q", "")
        ctx["selected_category"] = self.request.GET.get("category", "")
        ctx["selected_status"] = self.request.GET.get("status", "")
        return ctx


class ProductCreateView(SuperAdminRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = "dashboard/product_form.html"
    success_url = reverse_lazy("dashboard:products")

    def form_valid(self, form):
        response = super().form_valid(form)
        self._save_uploaded_images()
        messages.success(self.request, f'Product "{self.object.name}" was created successfully.')
        return response

    def _save_uploaded_images(self):
        files = self.request.FILES.getlist("images")
        for i, f in enumerate(files):
            ProductImage.objects.create(
                product=self.object,
                image=f,
                display_order=i,
                is_primary=(i == 0 and not self.object.images.exists()),
            )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["is_edit"] = False
        return ctx


class ProductUpdateView(SuperAdminRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = "dashboard/product_form.html"
    success_url = reverse_lazy("dashboard:products")

    def form_valid(self, form):
        response = super().form_valid(form)
        files = self.request.FILES.getlist("images")
        start = self.object.images.count()
        for i, f in enumerate(files):
            ProductImage.objects.create(
                product=self.object,
                image=f,
                display_order=start + i,
                is_primary=(start == 0 and i == 0),
            )
        messages.success(self.request, f'Product "{self.object.name}" was updated.')
        return response

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["is_edit"] = True
        ctx["existing_images"] = self.object.images.all()
        return ctx


class ProductDeleteView(SuperAdminRequiredMixin, DeleteView):
    model = Product
    success_url = reverse_lazy("dashboard:products")
    template_name = "dashboard/confirm_delete.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = f'Delete Product "{self.object.name}"?'
        ctx["cancel_url"] = reverse_lazy("dashboard:products")
        return ctx

    def form_valid(self, form):
        messages.success(self.request, f'Product "{self.object.name}" was deleted.')
        return super().form_valid(form)


class ProductTogglePublishView(SuperAdminRequiredMixin, View):
    def post(self, request, pk):
        product = get_object_or_404(Product, pk=pk)
        product.is_published = not product.is_published
        product.save(update_fields=["is_published"])
        messages.success(request, f'"{product.name}" status updated to {"Published" if product.is_published else "Draft"}.')
        return redirect("dashboard:products")


class ProductImageDeleteView(SuperAdminRequiredMixin, View):
    def post(self, request, pk, image_id):
        image = get_object_or_404(ProductImage, pk=image_id, product_id=pk)
        image.delete()
        messages.success(request, "Image removed.")
        return redirect("dashboard:product_edit", pk=pk)


class CategoryManagementView(SuperAdminRequiredMixin, ListView):
    model = Category
    template_name = "dashboard/categories.html"
    context_object_name = "categories"

    def get_queryset(self):
        return Category.objects.annotate(product_count=Count("products")).order_by("display_order", "name")


class CategoryCreateView(SuperAdminRequiredMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = "dashboard/category_form.html"
    success_url = reverse_lazy("dashboard:categories")

    def form_valid(self, form):
        messages.success(self.request, f'Category "{form.instance.name}" created.')
        return super().form_valid(form)


class CategoryUpdateView(SuperAdminRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = "dashboard/category_form.html"
    success_url = reverse_lazy("dashboard:categories")

    def form_valid(self, form):
        messages.success(self.request, f'Category "{form.instance.name}" updated.')
        return super().form_valid(form)


class CategoryDeleteView(SuperAdminRequiredMixin, DeleteView):
    model = Category
    success_url = reverse_lazy("dashboard:categories")
    template_name = "dashboard/confirm_delete.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = f'Delete Category "{self.object.name}"?'
        ctx["cancel_url"] = reverse_lazy("dashboard:categories")
        return ctx

    def form_valid(self, form):
        messages.success(self.request, f'Category "{self.object.name}" was deleted.')
        return super().form_valid(form)


class CollectionManagementView(SuperAdminRequiredMixin, ListView):
    model = Collection
    template_name = "dashboard/collections.html"
    context_object_name = "collections"

    def get_queryset(self):
        return Collection.objects.annotate(product_count=Count("products")).order_by("name")


class CollectionCreateView(SuperAdminRequiredMixin, CreateView):
    model = Collection
    fields = ["name", "slug", "description", "banner_image", "is_active"]
    template_name = "dashboard/collection_form.html"
    success_url = reverse_lazy("dashboard:collections")

    def form_valid(self, form):
        messages.success(self.request, f'Collection "{form.instance.name}" created.')
        return super().form_valid(form)


class CollectionUpdateView(SuperAdminRequiredMixin, UpdateView):
    model = Collection
    fields = ["name", "slug", "description", "banner_image", "is_active"]
    template_name = "dashboard/collection_form.html"
    success_url = reverse_lazy("dashboard:collections")

    def form_valid(self, form):
        messages.success(self.request, f'Collection "{form.instance.name}" updated.')
        return super().form_valid(form)


class CollectionDeleteView(SuperAdminRequiredMixin, DeleteView):
    model = Collection
    success_url = reverse_lazy("dashboard:collections")
    template_name = "dashboard/confirm_delete.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = f'Delete Collection "{self.object.name}"?'
        ctx["cancel_url"] = reverse_lazy("dashboard:collections")
        return ctx

    def form_valid(self, form):
        messages.success(self.request, f'Collection "{self.object.name}" deleted.')
        return super().form_valid(form)


class InventoryView(SuperAdminRequiredMixin, ListView):
    template_name = "dashboard/inventory.html"
    context_object_name = "products"
    paginate_by = 20

    def get_queryset(self):
        filter_type = self.request.GET.get("filter", "all")
        qs = Product.objects.all().select_related("category")
        if filter_type == "low_stock":
            qs = qs.filter(stock_quantity__lte=5, stock_quantity__gt=0)
        elif filter_type == "out_of_stock":
            qs = qs.filter(stock_quantity=0)
        return qs.order_by("stock_quantity")

    def post(self, request):
        product_id = request.POST.get("product_id")
        stock_quantity = request.POST.get("stock_quantity")
        if product_id and stock_quantity is not None:
            product = get_object_or_404(Product, pk=product_id)
            try:
                product.stock_quantity = max(0, int(stock_quantity))
                product.save(update_fields=["stock_quantity"])
                messages.success(request, f'Stock for "{product.name}" updated to {product.stock_quantity}.')
            except ValueError:
                messages.error(request, "Invalid stock value.")
        return redirect("dashboard:inventory")


class ReviewModerationView(SuperAdminRequiredMixin, ListView):
    model = Review
    template_name = "dashboard/reviews.html"
    context_object_name = "reviews"
    paginate_by = 15

    def get_queryset(self):
        return Review.objects.all().select_related("product", "user").order_by("-created_at")


class ReviewToggleApproveView(SuperAdminRequiredMixin, View):
    def post(self, request, pk):
        review = get_object_or_404(Review, pk=pk)
        review.is_approved = not review.is_approved
        review.save(update_fields=["is_approved"])
        messages.success(request, f'Review by {review.user} is now {"Approved" if review.is_approved else "Unapproved"}.')
        return redirect("dashboard:reviews")


class ReviewDeleteView(SuperAdminRequiredMixin, DeleteView):
    model = Review
    success_url = reverse_lazy("dashboard:reviews")
    template_name = "dashboard/confirm_delete.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = "Delete Product Review?"
        ctx["cancel_url"] = reverse_lazy("dashboard:reviews")
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Review deleted.")
        return super().form_valid(form)


# ---------------------------------------------------------------------------
# 3. Order Management (Orders, Payments, Refunds, Shipping)
# ---------------------------------------------------------------------------

class OrderManagementView(SuperAdminRequiredMixin, ListView):
    template_name = "dashboard/orders.html"
    context_object_name = "orders"
    paginate_by = 15

    def get_queryset(self):
        qs = Order.objects.all().select_related("user").prefetch_related("items", "payments")
        q = self.request.GET.get("q", "").strip()
        status = self.request.GET.get("status", "")

        if q:
            qs = qs.filter(Q(order_number__icontains=q) | Q(user__username__icontains=q) | Q(shipping_full_name__icontains=q))
        if status:
            qs = qs.filter(status=status)

        return qs.order_by("-placed_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["statuses"] = Order.Status.choices
        ctx["selected_status"] = self.request.GET.get("status", "")
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


class OrderDetailView(SuperAdminRequiredMixin, DetailView):
    model = Order
    template_name = "dashboard/order_detail.html"
    context_object_name = "order"

    def get_queryset(self):
        return Order.objects.select_related("user", "shipping_address", "billing_address").prefetch_related("items__product", "payments")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["statuses"] = Order.Status.choices
        return ctx


class OrderStatusUpdateView(SuperAdminRequiredMixin, View):
    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        new_status = request.POST.get("status")
        tracking_number = request.POST.get("tracking_number")

        if new_status in dict(Order.Status.choices):
            order.status = new_status
            if tracking_number is not None:
                order.tracking_number = tracking_number
            order.save()
            messages.success(request, f"Order #{order.order_number} status updated to {order.get_status_display()}.")
        return redirect("dashboard:order_detail", pk=pk)


class PaymentListView(SuperAdminRequiredMixin, ListView):
    model = Payment
    template_name = "dashboard/payments.html"
    context_object_name = "payments"
    paginate_by = 15

    def get_queryset(self):
        return Payment.objects.all().select_related("order__user").order_by("-created_at")


class VerifyPaymentView(SuperAdminRequiredMixin, View):
    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        payment = order.payments.order_by("-created_at").first()
        if payment:
            payment.status = Payment.Status.SUCCESS
            payment.save(update_fields=["status"])

        order.status = Order.Status.CONFIRMED
        order.save(update_fields=["status"])
        messages.success(request, f"Order #{order.order_number} payment verified & status marked Confirmed.")
        return redirect("dashboard:order_detail", pk=pk)


class RefundListView(SuperAdminRequiredMixin, ListView):
    model = ReturnRequest
    template_name = "dashboard/refunds.html"
    context_object_name = "refunds"
    paginate_by = 15

    def get_queryset(self):
        return ReturnRequest.objects.all().select_related("order_item__order__user", "order_item__product").order_by("-requested_at")


class ApproveRefundView(SuperAdminRequiredMixin, View):
    def post(self, request, pk):
        refund = get_object_or_404(ReturnRequest, pk=pk)
        action = request.POST.get("action")
        if action == "approve":
            refund.status = ReturnRequest.Status.APPROVED
            refund.admin_notes = "Approved by Super Admin."
            messages.success(request, f"Return request #{refund.id} approved.")
        elif action == "reject":
            refund.status = ReturnRequest.Status.REJECTED
            refund.admin_notes = "Rejected by Super Admin."
            messages.info(request, f"Return request #{refund.id} rejected.")
        refund.save()
        return redirect("dashboard:refunds")


# ---------------------------------------------------------------------------
# 4. Customer Management (Customers, Addresses, Wishlists)
# ---------------------------------------------------------------------------

class CustomerListView(SuperAdminRequiredMixin, ListView):
    template_name = "dashboard/customers.html"
    context_object_name = "customers"
    paginate_by = 15

    def get_queryset(self):
        qs = User.objects.filter(is_staff=False).annotate(
            total_orders=Count("orders"),
            total_spent=Sum("orders__grand_total")
        )
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(Q(username__icontains=q) | Q(email__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q))
        return qs.order_by("-date_joined")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["q"] = self.request.GET.get("q", "")
        return ctx


class CustomerDetailView(SuperAdminRequiredMixin, DetailView):
    model = User
    template_name = "dashboard/customer_detail.html"
    context_object_name = "customer"

    def get_queryset(self):
        return User.objects.filter(is_staff=False)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        customer = self.object
        ctx["addresses"] = Address.objects.filter(user=customer)
        ctx["orders"] = Order.objects.filter(user=customer).order_by("-placed_at")
        ctx["wishlist_items"] = Wishlist.objects.filter(user=customer).select_related("product")
        ctx["reviews"] = Review.objects.filter(user=customer).select_related("product")
        return ctx


class CustomerToggleActiveView(SuperAdminRequiredMixin, View):
    def post(self, request, pk):
        customer = get_object_or_404(User, pk=pk, is_staff=False)
        customer.is_active = not customer.is_active
        customer.save(update_fields=["is_active"])
        status_text = "activated" if customer.is_active else "deactivated/blocked"
        messages.success(request, f"Customer account {customer.username} is now {status_text}.")
        return redirect("dashboard:customer_detail", pk=pk)


# ---------------------------------------------------------------------------
# 5. Marketing (Banners, Coupons, Newsletter Subscribers)
# ---------------------------------------------------------------------------

class BannerListView(SuperAdminRequiredMixin, ListView):
    model = HeroBanner
    template_name = "dashboard/banners.html"
    context_object_name = "banners"

    def get_queryset(self):
        return HeroBanner.objects.all().order_by("display_order")


class BannerCreateView(SuperAdminRequiredMixin, CreateView):
    model = HeroBanner
    form_class = HeroBannerForm
    template_name = "dashboard/banner_form.html"
    success_url = reverse_lazy("dashboard:banners")

    def form_valid(self, form):
        messages.success(self.request, "Homepage Banner created.")
        return super().form_valid(form)


class BannerUpdateView(SuperAdminRequiredMixin, UpdateView):
    model = HeroBanner
    form_class = HeroBannerForm
    template_name = "dashboard/banner_form.html"
    success_url = reverse_lazy("dashboard:banners")

    def form_valid(self, form):
        messages.success(self.request, "Homepage Banner updated.")
        return super().form_valid(form)


class BannerDeleteView(SuperAdminRequiredMixin, DeleteView):
    model = HeroBanner
    success_url = reverse_lazy("dashboard:banners")
    template_name = "dashboard/confirm_delete.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = f'Delete Banner "{self.object.title}"?'
        ctx["cancel_url"] = reverse_lazy("dashboard:banners")
        return ctx

    def form_valid(self, form):
        messages.success(self.request, "Banner deleted.")
        return super().form_valid(form)


class CouponListView(SuperAdminRequiredMixin, ListView):
    model = Coupon
    template_name = "dashboard/coupons.html"
    context_object_name = "coupons"
    paginate_by = 15

    def get_queryset(self):
        return Coupon.objects.all().order_by("-valid_until")


class CouponCreateView(SuperAdminRequiredMixin, CreateView):
    model = Coupon
    fields = [
        "code", "description", "discount_type", "discount_value",
        "min_order_value", "max_discount_amount", "usage_limit",
        "user_limit", "valid_from", "valid_until", "is_active",
    ]
    template_name = "dashboard/coupon_form.html"
    success_url = reverse_lazy("dashboard:coupons")

    def form_valid(self, form):
        messages.success(self.request, f'Coupon "{form.instance.code}" created.')
        return super().form_valid(form)


class CouponUpdateView(SuperAdminRequiredMixin, UpdateView):
    model = Coupon
    fields = [
        "code", "description", "discount_type", "discount_value",
        "min_order_value", "max_discount_amount", "usage_limit",
        "user_limit", "valid_from", "valid_until", "is_active",
    ]
    template_name = "dashboard/coupon_form.html"
    success_url = reverse_lazy("dashboard:coupons")

    def form_valid(self, form):
        messages.success(self.request, f'Coupon "{form.instance.code}" updated.')
        return super().form_valid(form)


class CouponDeleteView(SuperAdminRequiredMixin, DeleteView):
    model = Coupon
    success_url = reverse_lazy("dashboard:coupons")
    template_name = "dashboard/confirm_delete.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["title"] = f'Delete Coupon "{self.object.code}"?'
        ctx["cancel_url"] = reverse_lazy("dashboard:coupons")
        return ctx

    def form_valid(self, form):
        messages.success(self.request, f'Coupon "{self.object.code}" deleted.')
        return super().form_valid(form)


class NewsletterSubscriberListView(SuperAdminRequiredMixin, TemplateView):
    template_name = "dashboard/newsletter.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["subscribers"] = User.objects.filter(is_staff=False, is_email_verified=True)
        return ctx


# ---------------------------------------------------------------------------
# 6. Analytics
# ---------------------------------------------------------------------------

class AnalyticsView(SuperAdminRequiredMixin, TemplateView):
    template_name = "dashboard/analytics.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        valid_statuses = [
            Order.Status.CONFIRMED,
            Order.Status.PROCESSING,
            Order.Status.SHIPPED,
            Order.Status.OUT_FOR_DELIVERY,
            Order.Status.DELIVERED,
        ]
        valid_orders = Order.objects.filter(status__in=valid_statuses)

        ctx["total_revenue"] = valid_orders.aggregate(total=Sum("grand_total"))["total"] or Decimal("0.00")
        ctx["total_lifetime_revenue"] = ctx["total_revenue"]
        ctx["total_orders_count"] = valid_orders.count()
        ctx["total_confirmed_orders"] = ctx["total_orders_count"]
        ctx["average_order_value"] = (
            ctx["total_revenue"] / ctx["total_orders_count"] if ctx["total_orders_count"] > 0 else Decimal("0.00")
        )

        ctx["best_selling_products"] = Product.objects.annotate(
            sold_count=Sum("orderitem__quantity"),
            total_sales_val=Sum(F("orderitem__unit_price_snapshot") * F("orderitem__quantity"))
        ).filter(sold_count__gt=0).order_by("-sold_count")[:10]

        ctx["top_categories"] = Category.objects.annotate(
            total_sales_val=Sum("products__orderitem__unit_price_snapshot")
        ).order_by("-total_sales_val")[:5]

        return ctx


class DailyRevenueApiView(SuperAdminRequiredMixin, View):
    def get(self, request):
        try:
            days = int(request.GET.get("days", 14))
        except (ValueError, TypeError):
            days = 14
        days = max(7, min(90, days))
        start_date = timezone.now().date() - timezone.timedelta(days=days - 1)

        valid_statuses = [
            Order.Status.CONFIRMED,
            Order.Status.PROCESSING,
            Order.Status.SHIPPED,
            Order.Status.OUT_FOR_DELIVERY,
            Order.Status.DELIVERED,
        ]

        orders_qs = Order.objects.filter(
            status__in=valid_statuses,
            placed_at__date__gte=start_date
        )

        revenue_by_day = dict(
            orders_qs.annotate(day=TruncDate("placed_at"))
            .values("day")
            .annotate(total=Sum("grand_total"))
            .values_list("day", "total")
        )

        labels = []
        date_strings = []
        revenues = []
        current_date = start_date
        today = timezone.now().date()

        while current_date <= today:
            val = float(revenue_by_day.get(current_date) or 0)
            labels.append(current_date.strftime("%b %d"))
            date_strings.append(current_date.strftime("%Y-%m-%d"))
            revenues.append(round(val, 2))
            current_date += timezone.timedelta(days=1)

        total_period_revenue = sum(revenues)
        has_data = len(revenues) > 0
        avg_val = total_period_revenue / len(revenues) if revenues else 0.0
        avg_rounded = round(avg_val, 2)
        average_line = [avg_rounded] * len(revenues)

        return JsonResponse({
            "has_data": has_data,
            "labels": labels,
            "date_strings": date_strings,
            "revenues": revenues,
            "average": avg_rounded,
            "average_daily_revenue": avg_rounded,
            "total_period_revenue": round(total_period_revenue, 2),
            "average_line": average_line,
        })


# ---------------------------------------------------------------------------
# 7. Reports (Excel Exports & PDF Summaries)
# ---------------------------------------------------------------------------

class ReportsOverviewView(SuperAdminRequiredMixin, TemplateView):
    template_name = "dashboard/reports.html"


class ReportExportOrdersExcelView(SuperAdminRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Orders Report")
        else:
            ws.title = "Orders Report"

        headers = ["Order Number", "Customer", "Date", "Status", "Items", "Grand Total (₹)", "Payment Method"]
        ws.append(headers)

        header_fill = PatternFill(start_color="141416", end_color="141416", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="C9A453")

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        orders = Order.objects.all().select_related("user").prefetch_related("items", "payments").order_by("-placed_at")
        for o in orders:
            latest_payment = o.payments.order_by("-created_at").first()
            payment_method = latest_payment.get_gateway_display() if latest_payment else "N/A"
            ws.append([
                o.order_number,
                o.user.get_full_name() or o.user.username,
                o.placed_at.strftime("%Y-%m-%d %H:%M"),
                o.get_status_display(),
                o.items.count(),
                float(o.grand_total),
                payment_method,
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="Eternaaura_Orders_Report.xlsx"'
        wb.save(response)
        return response


class ReportExportProductsExcelView(SuperAdminRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Products Inventory")
        else:
            ws.title = "Products Inventory"

        headers = ["SKU", "Product Name", "Category", "Price (₹)", "Stock", "Status"]
        ws.append(headers)

        header_fill = PatternFill(start_color="141416", end_color="141416", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="C9A453")

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        products = Product.objects.all().select_related("category").order_by("name")
        for p in products:
            ws.append([
                p.sku,
                p.name,
                p.category.name if p.category else "Uncategorized",
                float(p.price),
                p.stock_quantity,
                "Published" if p.is_published else "Draft",
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="Eternaaura_Products_Inventory.xlsx"'
        wb.save(response)
        return response


class ReportExportCustomersExcelView(SuperAdminRequiredMixin, View):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Customers Report")
        else:
            ws.title = "Customers Report"

        headers = ["Username", "Email", "Phone", "Joined Date", "Total Orders", "Total Spent (₹)"]
        ws.append(headers)

        header_fill = PatternFill(start_color="141416", end_color="141416", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="C9A453")

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        customers = User.objects.filter(is_staff=False).annotate(
            total_orders=Count("orders"),
            total_spent=Sum("orders__grand_total")
        ).order_by("-date_joined")

        for c in customers:
            ws.append([
                c.username,
                c.email,
                c.phone_number or "N/A",
                c.date_joined.strftime("%Y-%m-%d"),
                c.total_orders,
                float(c.total_spent or 0),
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="Eternaaura_Customers_Report.xlsx"'
        wb.save(response)
        return response


class ReportSalesPdfView(SuperAdminRequiredMixin, View):
    def get(self, request):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            leading=24,
            textColor=colors.HexColor('#8A6F28'),
            alignment=1
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=12,
            textColor=colors.HexColor('#555555'),
            alignment=1
        )

        elements.append(Paragraph("ETERNAAURA — Executive Sales Report", title_style))
        elements.append(Paragraph(f"Generated on {timezone.now().strftime('%d %B %Y, %H:%M')}", subtitle_style))
        elements.append(Spacer(1, 20))

        # Metrics Table
        total_products = Product.objects.count()
        total_customers = User.objects.filter(is_staff=False).count()
        total_orders = Order.objects.count()
        total_revenue = Order.objects.filter(status=Order.Status.DELIVERED).aggregate(t=Sum("grand_total"))["t"] or Decimal("0.00")

        data = [
            ["Metric", "Value"],
            ["Total Registered Customers", str(total_customers)],
            ["Total Active Products", str(total_products)],
            ["Total Orders Placed", str(total_orders)],
            ["Total Delivered Sales Revenue", f"INR {total_revenue:,.2f}"],
        ]

        t = Table(data, colWidths=[250, 250])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#141416')),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.HexColor('#C9A453')),
            ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
        ]))
        elements.append(t)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="Eternaaura_Sales_Summary.pdf"'
        return response


# ---------------------------------------------------------------------------
# 8. Website Settings (Store Settings, Delivery, Taxes, Payment Gateways)
# ---------------------------------------------------------------------------

class StoreSettingsView(SuperAdminRequiredMixin, View):
    template_name = "dashboard/settings.html"

    def get(self, request):
        settings_obj = StoreSetting.get_solo()
        return render(request, self.template_name, {"settings": settings_obj})

    def post(self, request):
        settings_obj = StoreSetting.get_solo()
        settings_obj.store_name = request.POST.get("store_name", settings_obj.store_name)
        settings_obj.contact_email = request.POST.get("contact_email", settings_obj.contact_email)
        settings_obj.support_phone = request.POST.get("support_phone", settings_obj.support_phone)
        settings_obj.store_address = request.POST.get("store_address", settings_obj.store_address)
        settings_obj.currency_symbol = request.POST.get("currency_symbol", settings_obj.currency_symbol)

        try:
            settings_obj.tax_percentage = Decimal(request.POST.get("tax_percentage", "3.00"))
            settings_obj.standard_shipping_fee = Decimal(request.POST.get("standard_shipping_fee", "150.00"))
            settings_obj.free_shipping_threshold = Decimal(request.POST.get("free_shipping_threshold", "5000.00"))
        except (ValueError, ArithmeticError):
            messages.error(request, "Invalid numeric settings input.")

        settings_obj.enable_razorpay = request.POST.get("enable_razorpay") == "on"
        settings_obj.enable_stripe = request.POST.get("enable_stripe") == "on"
        settings_obj.enable_cod = request.POST.get("enable_cod") == "on"
        settings_obj.enable_upi = request.POST.get("enable_upi") == "on"

        settings_obj.save()
        messages.success(request, "Store settings updated successfully.")
        return redirect("dashboard:settings")


# ---------------------------------------------------------------------------
# 9. Admin Profile
# ---------------------------------------------------------------------------

class AdminProfileView(SuperAdminRequiredMixin, View):
    template_name = "dashboard/profile.html"

    def get(self, request):
        password_form = PasswordChangeForm(request.user)
        return render(request, self.template_name, {
            "password_form": password_form
        })

    def post(self, request):
        action = request.POST.get("action")
        if action == "update_info":
            request.user.first_name = request.POST.get("first_name", "")
            request.user.last_name = request.POST.get("last_name", "")
            request.user.email = request.POST.get("email", request.user.email)
            request.user.phone_number = request.POST.get("phone_number", request.user.phone_number)
            request.user.save()
            messages.success(request, "Admin profile info updated.")
            return redirect("dashboard:profile")

        elif action == "change_password":
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Password changed successfully.")
                return redirect("dashboard:profile")
            else:
                messages.error(request, "Please correct the password errors below.")
                return render(request, self.template_name, {"password_form": password_form})

        return redirect("dashboard:profile")
